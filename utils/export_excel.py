"""
导出文件
"""
from io import BytesIO

from openpyxl import Workbook

# workbook = Workbook()
#
# # 默认sheet
# sheet = workbook.active
# sheet.title = "sheet1"
# # sheet = workbook.create_sheet(title="新sheet")
columns = ['1', '2', '3']
datas = [
    [1, 2, 3],
    [1, 2, 3],
    [1, 2, 3],
]


# sheet.append(columns)
# for data in datas:
#     sheet.append(data)
#
# workbook.save('./111.xlsx')


def export(columns, datas):
    workbook = Workbook()
    # 默认sheet
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(columns)
    for d in datas:
        sheet.append(d)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


if __name__ == '__main__':
    print(export(columns, datas))
